import streamlit as st
import pandas as pd
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from copy import copy

# ==========================================
# CONFIGURACIÓ: Noms exactes de les columnes
# ==========================================
COL_FECHA_PEST = "DATA "
COL_CLIENTE_PEST = "CLIENT A FACTURAR"
COL_IMPORTE_PEST = "IMPORT"
COL_FACTURA_PEST = "FACTURA"
COL_REBUT_PEST = "REBUT"
COL_CONCEPTE_PEST = "CONCEPTE"

COL_FECHA_PRIN = "Data pagament"
COL_CLIENTE_PRIN = "Nom"
COL_IMPORTE_PRIN = "Import"
# ==========================================

# --- Estils CSS per millorar l'aspecte sense canviar l'estructura ---
st.set_page_config(page_title="Comparador Financer", layout="wide")
st.markdown("""
    <style>
    /* Fons sidebar més suau */
    [data-testid="stSidebar"] > div:first-child { background-color: #f8f9fa; }
    
    /* Metrics amb estil de targeta */
    div[data-testid="stMetricValue"] { font-size: 1.8rem; font-weight: bold; color: #1E3A8A; }
    div[data-testid="stMetricLabel"] { font-size: 1rem; font-weight: 600; }
    [data-testid="stMetric"] { background-color: #ffffff; border: 1px solid #e5e7eb; border-radius: 8px; padding: 15px 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }
    
    /* Expansors més definits */
    .stExpander { border: 1px solid #e5e7eb; border-radius: 8px; margin-bottom: 10px; }
    streamlit-expanderHeader { font-weight: bold; font-size: 1.1rem; }
    </style>
""", unsafe_allow_html=True)

def netejar_nom_empresa(nom):
    if pd.isna(nom):
        return ""
    nom = str(nom).upper()
    nom = re.sub(r'[.,]', '', nom)
    nom = re.sub(r'\s(SA|SL|SAB|SLU|S\s*A|S\s*L)\s*$', '', nom)
    nom = re.sub(r'\s+', ' ', nom).strip()
    return nom

@st.cache_data
def processar_dades(archivo_principal_bytes, archivo_pestana_bytes, hojas_seleccionadas):
    df_principal = pd.read_excel(io.BytesIO(archivo_principal_bytes))
    lista_df = []
    xls = pd.ExcelFile(io.BytesIO(archivo_pestana_bytes))
    
    for nombre_hoja in xls.sheet_names:
        if nombre_hoja in hojas_seleccionadas:
            df_hoja = pd.read_excel(xls, sheet_name=nombre_hoja)
            df_hoja['Origen_Pestana'] = nombre_hoja
            df_hoja['Fila_Excel'] = df_hoja.index + 2 
            lista_df.append(df_hoja)
            
    df_pestana = pd.concat(lista_df, ignore_index=True)

    # Neteja
    df_pestana['Fecha_Limpia'] = pd.to_datetime(df_pestana[COL_FECHA_PEST], errors='coerce', dayfirst=True).dt.strftime('%Y-%m-%d').fillna('SENSE_DATA')
    df_principal['Fecha_Limpia'] = pd.to_datetime(df_principal[COL_FECHA_PRIN], errors='coerce', dayfirst=True).dt.strftime('%Y-%m-%d').fillna('SENSE_DATA')
    df_pestana['Nombre_Match'] = df_pestana[COL_CLIENTE_PEST].apply(netejar_nom_empresa)
    df_principal['Nombre_Match'] = df_principal[COL_CLIENTE_PRIN].apply(netejar_nom_empresa)
    df_pestana['Importe_Str'] = df_pestana[COL_IMPORTE_PEST].astype(str).str.replace(',', '.')
    df_principal['Importe_Str'] = df_principal[COL_IMPORTE_PRIN].astype(str).str.replace(',', '.')
    df_pestana['Importe_Limpio'] = pd.to_numeric(df_pestana['Importe_Str'], errors='coerce').map(lambda x: f"{x:.2f}" if pd.notnull(x) else "SENSE_IMPORT")
    df_principal['Importe_Limpio'] = pd.to_numeric(df_principal['Importe_Str'], errors='coerce').map(lambda x: f"{x:.2f}" if pd.notnull(x) else "SENSE_IMPORT")
    df_pestana['Concepte_Check'] = df_pestana[COL_CONCEPTE_PEST].fillna('').astype(str).str.upper().str.strip()

    # Cerca 1: Pestanyes -> Principal
    def check_pestana_en_principal(row, df_prin):
        nombre_pest = row['Nombre_Match']
        cand_fecha_imp = df_prin[(df_prin['Fecha_Limpia'] == row['Fecha_Limpia']) & (df_prin['Importe_Limpio'] == row['Importe_Limpio'])]
        if not cand_fecha_imp.empty:
            for _, row_prin in cand_fecha_imp.iterrows():
                nombre_prin = row_prin['Nombre_Match']
                if nombre_prin != "" and nombre_pest != "" and (nombre_prin in nombre_pest or nombre_pest in nombre_prin):
                    return 'TROBAT'
            return 'FALTA'
        return 'FALTA'

    df_pestana['Estat_al_Principal'] = df_pestana.apply(lambda row: check_pestana_en_principal(row, df_principal), axis=1)
    
    # Cerca 2: Principal -> Pestanyes 
    def check_principal_en_pestanyes(row, df_pest):
        nombre_prin = row['Nombre_Match']
        cand_fecha_imp = df_pest[(df_pest['Fecha_Limpia'] == row['Fecha_Limpia']) & (df_pest['Importe_Limpio'] == row['Importe_Limpio'])]
        if not cand_fecha_imp.empty:
            for _, row_pest in cand_fecha_imp.iterrows():
                nombre_pest = row_pest['Nombre_Match']
                if nombre_prin != "" and nombre_pest != "" and (nombre_prin in nombre_pest or nombre_pest in nombre_prin):
                    return 'TROBAT'
            return 'FALTA'
        return 'FALTA'

    df_principal['Estat_a_Pestanes'] = df_principal.apply(lambda row: check_principal_en_pestanyes(row, df_pestana), axis=1)
    df_principal['Fila_Excel'] = df_principal.index + 2 

    # Comprovacions de qualitat
    df_pestana['Es_Repetit'] = df_pestana.duplicated(subset=['Fecha_Limpia', 'Importe_Limpio', 'Nombre_Match', 'Concepte_Check'], keep=False).map({True: 'SI REPETIT', False: 'NO'})
    df_pestana['Factura_Check'] = df_pestana[COL_FACTURA_PEST].fillna('').astype(str).str.strip()
    df_pestana['Rebut_Check'] = df_pestana[COL_REBUT_PEST].fillna('').astype(str).str.strip()
    df_pestana['Estat_Documentacio'] = 'OK'
    mask_no_doc = (df_pestana['Factura_Check'] == '') & (df_pestana['Rebut_Check'] == '')
    df_pestana.loc[mask_no_doc, 'Estat_Documentacio'] = 'FALTA FACTURA I REBUT'

    # Separar resultats
    df_faltan_pest = df_pestana[df_pestana['Estat_al_Principal'] == 'FALTA'].copy()
    df_faltan_prin = df_principal[df_principal['Estat_a_Pestanes'] == 'FALTA'].copy()
    df_repetidos = df_pestana[df_pestana['Es_Repetit'] == 'SI REPETIT'].copy()
    df_falta_doc = df_pestana[df_pestana['Estat_Documentacio'] == 'FALTA FACTURA I REBUT'].copy()

    return df_principal, df_pestana, df_faltan_prin, df_faltan_pest, df_repetidos, df_falta_doc

def apply_excel_formatting(original_file_bytes, output_df_bytes, num_original_cols):
    try:
        wb_orig = load_workbook(io.BytesIO(original_file_bytes))
        ws_orig = wb_orig.active
        orig_header_styles, orig_col_widths = {}, {}
        for cell in ws_orig[1]:
            if cell.value is not None:
                orig_header_styles[cell.column] = {'font': copy(cell.font), 'fill': copy(cell.fill), 'border': copy(cell.border), 'alignment': copy(cell.alignment)}
                orig_col_widths[cell.column] = ws_orig.column_dimensions[cell.column_letter].width

        wb_out = load_workbook(io.BytesIO(output_df_bytes))
        new_font, new_fill = Font(bold=True, color="FFFFFF"), PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        new_border, new_alignment = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for ws_out in wb_out.worksheets:
            for cell in ws_out[1]:
                cell.border = new_border
                if cell.column <= num_original_cols and cell.column in orig_header_styles:
                    style = orig_header_styles[cell.column]
                    cell.font, cell.fill, cell.border, cell.alignment = style['font'], style['fill'], style['border'], style['alignment']
                elif cell.column > num_original_cols:
                    cell.font, cell.fill, cell.alignment = new_font, new_fill, new_alignment
            for row in ws_out.iter_rows(min_row=2, max_row=ws_out.max_row, max_col=ws_out.max_column):
                for cell in row:
                    cell.border = data_border
                    if cell.column <= num_original_cols and cell.column in orig_col_widths: ws_out.column_dimensions[cell.column_letter].width = orig_col_widths[cell.column]
                    elif cell.column > num_original_cols: ws_out.column_dimensions[cell.column_letter].width = 25

        final_output = io.BytesIO()
        wb_out.save(final_output)
        final_output.seek(0)
        return final_output
    except Exception as e:
        print(f"Error aplicant format: {e}")
        return io.BytesIO(output_df_bytes)

# ==========================================
# INTERFÍCIE ORIGINAL (VISUALMENT MILLORADA)
# ==========================================

st.title("Comparador")

st.sidebar.header("📂 Pujar Fitxers")
archivo_principal = st.sidebar.file_uploader("1️⃣ Excel LLISTAT DE COBRAMENTS", type=['xlsx', 'xls'])
archivo_pestana = st.sidebar.file_uploader("2️⃣ Excel INGRESSOS", type=['xlsx', 'xls'])

hojas_seleccionadas = []
if archivo_pestana:
    xls = pd.ExcelFile(archivo_pestana)
    hojas_seleccionadas = st.sidebar.multiselect("📑 Selecciona les pestanyes a comprovar:", options=xls.sheet_names, default=xls.sheet_names[:3])

if archivo_principal and archivo_pestana and hojas_seleccionadas:
    
    if st.sidebar.button("🚀 Iniciar Comparació", type="primary", use_container_width=True):
        with st.spinner("⏳ Processant i creuant dades... Aquesta operació pot trigar uns segons."):
            df_principal, df_pestana, df_faltan_prin, df_faltan_pest, df_repetidos, df_falta_doc = processar_dades(archivo_principal.getvalue(), archivo_pestana.getvalue(), hojas_seleccionadas)

            cols_tec = ['Fecha_Limpia', 'Nombre_Match', 'Importe_Str', 'Importe_Limpio', 'Factura_Check', 'Rebut_Check', 'Concepte_Check', 'Estat_a_Pestanes', 'Estat_al_Principal']
            cols_ocult = ['INCIDÈNCIES SECRETARIA', 'INCIDÈNCIES RECEPCIÓ MOSTRES', 'INCIDÈNCIES']
            num_original_cols = len(pd.read_excel(archivo_pestana, sheet_name=hojas_seleccionadas[0], nrows=0).columns)
            
            raw_output = io.BytesIO()
            with pd.ExcelWriter(raw_output, engine='openpyxl') as writer:
                df_pestana.drop(columns=cols_tec, errors='ignore').to_excel(writer, sheet_name='Tot_Creuat', index=False)
                df_faltan_pest.drop(columns=cols_tec + cols_ocult, errors='ignore').to_excel(writer, sheet_name='Falten_Pestanyes_al_Principal', index=False)
                df_faltan_prin.drop(columns=cols_tec, errors='ignore').to_excel(writer, sheet_name='Falten_Principal_a_Pestanyes', index=False)
                df_repetidos.drop(columns=cols_tec, errors='ignore').to_excel(writer, sheet_name='Repetits_a_Pestanyes', index=False)
                df_falta_doc.drop(columns=cols_tec, errors='ignore').to_excel(writer, sheet_name='Falta_Documentacio', index=False)
            raw_output.seek(0)
            final_output = apply_excel_formatting(archivo_pestana.getvalue(), raw_output.getvalue(), num_original_cols)

            st.session_state['resultats'] = {
                'df_principal': df_principal, 'df_pestana': df_pestana,
                'df_faltan_prin': df_faltan_prin, 'df_faltan_pest': df_faltan_pest,
                'df_repetidos': df_repetidos, 'df_falta_doc': df_falta_doc,
                'final_output': final_output,
                'cols_tec': cols_tec, 'cols_ocult': cols_ocult
            }

    if 'resultats' in st.session_state:
        res = st.session_state['resultats']
        df_principal = res['df_principal']
        df_pestana = res['df_pestana']
        df_faltan_prin = res['df_faltan_prin']
        df_faltan_pest = res['df_faltan_pest']
        df_repetidos = res['df_repetidos']
        df_falta_doc = res['df_falta_doc']
        
        st.success("✅ Procés finalitzat amb èxit.")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("🔴 FALTEN Cobraments", len(df_faltan_prin), delta="Ingressos sense registre", delta_color="inverse")
        col2.metric("🟠 FALTEN Ingressos", len(df_faltan_pest), delta="Cobraments sense ingrés", delta_color="inverse")
        col3.metric("🟡 REPETITS", len(df_repetidos))
        col4.metric("🔵 Falta FACT/REBUT", len(df_falta_doc))

        st.download_button("⬇️ Descarregar Excel de Resultats", data=res['final_output'], file_name="Comparacio_Resultat.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
        
        # 1. FALTANTS PRINCIPAL -> PESTANYES
        with st.expander("⬇️ Revisar FALTEN (Estan a Ingressos però no a Cobraments)", expanded=False):
            if df_faltan_prin.empty:
                st.info("✅ Tots els cobraments de Ingressos estan registrats al Listat de Cobraments.")
            else:
                cols_amagar_prin_pantalla = res['cols_tec'] + ['Estat_a_Pestanes']
                df_faltan_prin_pantalla = df_faltan_prin.drop(columns=cols_amagar_prin_pantalla, errors='ignore')
                st.dataframe(df_faltan_prin_pantalla, use_container_width=True, hide_index=True, height=300)
                
                # DESPLEGABLE INTERN PARA COMPARAR (OCULTO POR DEFECTO)
                with st.expander("🔍 Comparar fila concreta", expanded=False):
                    st.markdown("**Selecciona la fila del Principal que vols buscar a les Pestanyes:**")
                    
                    opcions_prin = []
                    for idx, row in df_faltan_prin.iterrows():
                        fila = int(row['Fila_Excel'])
                        client = str(row[COL_CLIENTE_PRIN])
                        import_ = str(row[COL_IMPORTE_PRIN])
                        opcions_prin.append(f"Principal Fila {fila} | {client} | {import_}€")
                    
                    seleccionat_prin = st.selectbox("Tria una fila:", options=opcions_prin, index=0, key="sel_prin")
                    
                    if seleccionat_prin:
                        idx_sel_prin = opcions_prin.index(seleccionat_prin)
                        original_row_prin = df_faltan_prin.iloc[idx_sel_prin]
                        
                        # Millora visual: Side-by-side comparison
                        col_orig, col_match = st.columns(2)
                        
                        with col_orig:
                            st.markdown("**📌 Fila Original:**")
                            df_orig_prin_display = pd.DataFrame([original_row_prin]).drop(columns=cols_amagar_prin_pantalla, errors='ignore')
                            st.dataframe(df_orig_prin_display.style.set_properties(**{'background-color': '#e7f1ff'}), use_container_width=True, hide_index=True)
                        
                        with col_match:
                            st.markdown("**🟩 Possibles Coincidències:**")
                            match_fi_pest = df_pestana[(df_pestana['Fecha_Limpia'] == original_row_prin['Fecha_Limpia']) & (df_pestana['Importe_Limpio'] == original_row_prin['Importe_Limpio'])]
                            cols_amagar_pest = res['cols_tec'] + ['Estat_al_Principal', 'Estat_Documentacio']
                            
                            if not match_fi_pest.empty:
                                match_pest_display = match_fi_pest.drop(columns=cols_amagar_pest, errors='ignore')
                                # Verd per coincidència exacte (data i import)
                                st.dataframe(match_pest_display.style.set_properties(**{'background-color': '#d4edda'}), use_container_width=True, hide_index=True)
                            else:
                                match_si_pest = df_pestana[(df_pestana['Importe_Limpio'] == original_row_prin['Importe_Limpio']) & (df_pestana['Fecha_Limpia'] != original_row_prin['Fecha_Limpia'])]
                                if not match_si_pest.empty:
                                    match_pest_display = match_si_pest.drop(columns=cols_amagar_pest, errors='ignore')
                                    # Groc per coincidència parcial (import però no data)
                                    st.warning("⚠️ Coincidència parcial: Mateix import, diferent data.")
                                    st.dataframe(match_pest_display.style.set_properties(**{'background-color': '#fff3cd'}), use_container_width=True, hide_index=True)
                                else:
                                    st.error("❌ No hi ha coincidències possibles a les pestanyes per aquest cobrament bancari.")

        # 2. FALTANTS PESTANYES -> PRINCIPAL
        with st.expander("⬆️ Revisar FALTEN (Estan a Cobraments però no a Ingressos)", expanded=False):
            if df_faltan_pest.empty:
                st.info("✅ Tots els cobraments interns estan a Ingressos.")
            else:
                cols_amagar_pantalla = res['cols_tec'] + res['cols_ocult'] + ['Estat_al_Principal', 'Es_Repetit', 'Estat_Documentacio']
                df_faltan_pest_pantalla = df_faltan_pest.drop(columns=cols_amagar_pantalla, errors='ignore')
                st.dataframe(df_faltan_pest_pantalla, use_container_width=True, hide_index=True, height=300)
                
                # DESPLEGABLE INTERN PARA COMPARAR (OCULTO POR DEFECTO)
                with st.expander("🔍 Comparar fila concreta", expanded=False):
                    st.markdown("**Selecciona la fila que vols comparar amb el Principal:**")
                    
                    opcions = []
                    for idx, row in df_faltan_pest.iterrows():
                        pestanya = str(row['Origen_Pestana'])
                        fila = int(row['Fila_Excel'])
                        client = str(row[COL_CLIENTE_PEST])
                        import_ = str(row[COL_IMPORTE_PEST])
                        opcions.append(f"{pestanya} | Fila {fila} | {client} | {import_}€")
                    
                    seleccionat = st.selectbox("Tria una fila:", options=opcions, index=0, key="sel_pest")
                    
                    if seleccionat:
                        idx_sel = opcions.index(seleccionat)
                        original_row = df_faltan_pest.iloc[idx_sel]
                        
                        # Millora visual: Side-by-side comparison
                        col_orig2, col_match2 = st.columns(2)
                        
                        with col_orig2:
                            st.markdown("**📌 Fila Original:**")
                            df_orig_display = pd.DataFrame([original_row]).drop(columns=cols_amagar_pantalla, errors='ignore')
                            st.dataframe(df_orig_display.style.set_properties(**{'background-color': '#fff3cd'}), use_container_width=True, hide_index=True)
                        
                        with col_match2:
                            st.markdown("**🟩 Possibles Coincidències:**")
                            match_fi = df_principal[(df_principal['Fecha_Limpia'] == original_row['Fecha_Limpia']) & (df_principal['Importe_Limpio'] == original_row['Importe_Limpio'])]
                            cols_amagar_prin = res['cols_tec'] + ['Estat_a_Pestanes']
                            
                            if not match_fi.empty:
                                match_display = match_fi.drop(columns=cols_amagar_prin, errors='ignore')
                                st.dataframe(match_display.style.set_properties(**{'background-color': '#d4edda'}), use_container_width=True, hide_index=True)
                            else:
                                match_si = df_principal[(df_principal['Importe_Limpio'] == original_row['Importe_Limpio']) & (df_principal['Fecha_Limpia'] != original_row['Fecha_Limpia'])]
                                if not match_si.empty:
                                    match_display = match_si.drop(columns=cols_amagar_prin, errors='ignore')
                                    st.warning("⚠️ Coincidència parcial: Mateix import, diferent data.")
                                    st.dataframe(match_display.style.set_properties(**{'background-color': '#fff3cd'}), use_container_width=True, hide_index=True)
                                else:
                                    st.error("❌ No hi ha coincidències possibles al principal per a aquesta fila.")

        # 3. REPETITS
        with st.expander("🔄 Revisar REPETITS"):
            if df_repetidos.empty:
                st.info("✅ No hi ha files repetides a les pestanyes.")
            else:
                st.dataframe(df_repetidos.drop(columns=res['cols_tec'], errors='ignore'), use_container_width=True, hide_index=True, height=300)

        # 4. FALTA DOCUMENTACIÓ
        with st.expander("📄 Revisar files sense FACTURA/REBUT"):
            if df_falta_doc.empty:
                st.info("✅ Tots els cobraments tenen o bé Factura o bé Rebut.")
            else:
                cols_amagar_doc = res['cols_tec'] + ['Estat_al_Principal', 'Es_Repetit', 'Estat_Documentacio'] + res['cols_ocult']
                df_falta_doc_pantalla = df_falta_doc.drop(columns=cols_amagar_doc, errors='ignore')
                st.dataframe(df_falta_doc_pantalla.style.set_properties(**{'background-color': '#ffe6e6'}), use_container_width=True, hide_index=True, height=300)

elif not archivo_principal or not archivo_pestana:
    st.info("Si us plau, puja els dos fitxers Excel a la barra lateral per començar.")
elif not hojas_seleccionadas:
    st.warning("⚠️ Si us plau, selecciona almenys una pestanya a la barra lateral.")